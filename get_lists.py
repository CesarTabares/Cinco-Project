# -*- coding: utf-8 -*-
"""
Created on Wed Jan 15 21:33:25 2020

@author: Cesar
"""
from webscraping import get_the_web
import time
import openpyxl


# Este proceso debe hacerse dentro de los chequeos periodicos para garantizar que no se hayan añadido nuevas ciudades o destinos
def get_cities_entities_web():
    
    wb_lists=openpyxl.load_workbook('Lists.xlsx')
    cities_entities_sheet=wb_lists['Cities-Entities']
    row_cell=1
    
    browser=get_the_web()
    obj_ciudad=(browser.find_element_by_id('ddlCiudad')).find_elements_by_tag_name('option')
    ciudades=[]
    
    
    for i in obj_ciudad:
        if i.get_attribute('value') != '0':
            ciudades.append(i.text)

    
    lista_prueba=[]
    for i in range(len(ciudades)):
        browser.find_element_by_id('ddlCiudad').send_keys(ciudades[i])
        time.sleep(2)
        obj_entidad=(browser.find_element_by_id('ddlEntidadEspecialidad')).find_elements_by_tag_name('option')

        for j in range(len(obj_entidad)):
            if obj_entidad[j].get_attribute('value') !='0':
                row_cell += 1
                cities_entities_sheet.cell(row=row_cell,column=1).value=ciudades[i]
                lista_prueba.append(obj_entidad[j].text)
                cities_entities_sheet.cell(row=row_cell,column=2).value=obj_entidad[j].text
            
    wb_lists.save('Lists.xlsx')
    
    
    print('done')
    
def make_cities_entities_dictionary():

    wb_lists=openpyxl.load_workbook('Lists.xlsx')
    cities_entities_sheet=wb_lists['Cities-Entities']
    cantidad_celdas=(len(cities_entities_sheet['A']))
    
    diccionario={}
    
    ciudades_choices=[]
    entidades_choices=[]
    
    for cell in cities_entities_sheet['A']:
        ciudades_choices.append(cell.value)
        
    for cell in cities_entities_sheet['B']:
        entidades_choices.append(cell.value)
    
    entidades_ciudad=[]
    ciudades_no_repetidas= sorted(set(ciudades_choices))

    for i in range(2,cantidad_celdas):
        
        if i == (cantidad_celdas-1):
            entidades_ciudad.append(entidades_choices[i])
            diccionario[ciudades_choices[i]]=entidades_ciudad
            break
        
        elif ciudades_choices[i]==ciudades_choices[i+1]:
            entidades_ciudad.append(entidades_choices[i])
    
        else:
             entidades_ciudad.append(entidades_choices[i])
             diccionario[ciudades_choices[i]]=entidades_ciudad
             entidades_ciudad=[]
    
    return diccionario,ciudades_no_repetidas

def make_others_list():
    wb_lists=openpyxl.load_workbook('Lists.xlsx')
    other_lists_sheet=wb_lists['Other-Lists']
    
    tipos_sujeto=[]
    tipos_persona=[]
    tipos_proceso=[]
    tipos_proceso_general=[]
    tipos_estado=[]

    for cell in other_lists_sheet['A']:
        if cell.value != None:
            tipos_sujeto.append(cell.value)

    for cell in other_lists_sheet['B']:
        if cell.value != None:
            tipos_persona.append(cell.value)
        
    for cell in other_lists_sheet['C']:
        if cell.value != None:
            tipos_proceso.append(cell.value)
            
    for cell in other_lists_sheet['D']:
        if cell.value != None:
            tipos_proceso_general.append(cell.value)
    
    for cell in other_lists_sheet['E']:
        if cell.value != None:
            tipos_estado.append(cell.value)
    
    smmlv= other_lists_sheet.cell(row=2,column=6).value

    return tipos_sujeto,tipos_persona,tipos_proceso, tipos_proceso_general, smmlv, tipos_estado
        
def get_clients_info(db_sheet,estado_abierto):
    
    
    list_id_clients=[]
    list_estado=[]
    list_open_clients=[]
    
    for cell in db_sheet['X']:
        list_id_clients.append(cell.value)
    
    for cell in db_sheet['Z']:
        list_estado.append(cell.value)
    
    index_clientes_abiertos=[i for i, value in enumerate(list_id_clients) if list_estado[i]==estado_abierto]

    for i in index_clientes_abiertos:
        list_open_clients.append(list_id_clients[i])
    
    list_open_clients=sorted(list(set(list_open_clients)))

    return list_open_clients,list_id_clients,list_estado

def get_client_process_open(db_sheet,estado_abierto,list_id_clients,list_estado,client):

    lista_procesos_abierto_cliente=[]
    lista_radicado_ini=[]
    
    for cell in db_sheet['B']:
        lista_radicado_ini.append(cell.value)
    
    index_procesos_abiertos=[i for i, value in enumerate(list_id_clients) if list_estado[i]==estado_abierto and list_id_clients[i]==client]   

    for i in index_procesos_abiertos:
        lista_procesos_abierto_cliente.append(lista_radicado_ini[i])
    
    return lista_procesos_abierto_cliente, lista_radicado_ini

def get_actuaciones_process_open(act_sheet,estado_abierto,radicado_ini):
    
    
    lista_radicado_ini_actsheet=[]
    lista_actuaciones_actsheet=[]
    lista_estados_actsheet=[]
    
    lista_actuaciones_proceso=[]
    
    for cell in act_sheet['B']:
        lista_radicado_ini_actsheet.append(cell.value)
        
    for cell in act_sheet['D']:
        lista_actuaciones_actsheet.append(cell.value)
        
    for cell in act_sheet['I']:
        lista_estados_actsheet.append(cell.value)
    
    index_actuaciones_abiertas_proceso=[i for i, value in enumerate(lista_actuaciones_actsheet) if lista_estados_actsheet[i]==estado_abierto and lista_radicado_ini_actsheet[i]==radicado_ini]
    
    for i in index_actuaciones_abiertas_proceso:
        lista_actuaciones_proceso.append(lista_actuaciones_actsheet[i])
    
    return lista_actuaciones_proceso,index_actuaciones_abiertas_proceso
    
 
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    