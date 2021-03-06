# -*- coding: utf-8 -*-
"""
Created on Thu Nov 28 14:28:41 2019

@author: user
"""
from selenium import webdriver
import wx
import time
import openpyxl
import os
from webscraping import asignar_nro_proceso, get_the_web, encontrar_actuaciones
from get_lists import get_cities_entities_web, make_cities_entities_dictionary, make_others_list, get_clients_info, get_client_process_open,get_actuaciones_process_open
from pubsub import pub

col_radicado_ini=2
col_radicado_completo=3
col_fecha_radicacion=4
col_tipo_general_proceso=5
col_tipo_especifico_proceso=6
col_cuantia=7
col_instancia=8
col_responsable=9
col_apoderado=10
col_ciudad=11
col_entidad=12
col_jurisdiccion=13
col_tipo_sujeto_cliente=14
col_tipo_persona_demandante=15
col_razon_social_demandante=16
col_nit_demandate=17
col_tipo_persona_demandado=18
col_razon_social_demandado=19
col_nit_demandado=20
col_tipo_persona_tercero=21
col_razon_social_tercero=22
col_nit_tercero=23
col_nit_cliente=24
col_nombre_cliente=25

#-- Excel BD Actuaciones--#
col_id_actuacion=1
col_numero_proceso=2
col_radicado_ini_act=3
col_fecha_actuacion=4
col_actuacion=5
col_anotacion=6
col_fecha_ini_termino=7
col_fecha_fin_termino=8
col_fecha_registro=9
col_estado=10
col_grupo=11
col_principal=12
col_actuacion_propia=13
estado_choices=['Abierto','Cerrado']
#-- Excel BD Actuaciones--#

#-- Excel Usuarios--#
col_nombre=1
col_usuario=2
col_correo=3
col_rol=4
col_tipo_usuario=5


#-- Excel Usuarios--#

DB = openpyxl.load_workbook('Database-Process.xlsx')
sheet = DB['Hoja1']


class MyFrame(wx.Frame):
    
    
    def OnKeyDown(self, event):
        """quit if user press q or Esc"""
        if event.GetKeyCode() == 27 or event.GetKeyCode() == ord('Q'): #27 is Esc
            self.Close(force=True)
            
        else:
            event.Skip()
 
    def __init__(self):
        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Software Legal", size=(1200, 700))  
        self.Bind(wx.EVT_KEY_UP, self.OnKeyDown)
        
        try:
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(image_file, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            self.panel = wx.StaticBitmap(self, -1, bmp1, (0, 0))
            
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
            
       
        
        
        button = wx.Button(self.panel, id=wx.ID_ANY, label="Ingresar Proceso" ,pos=(900, 100), size=(200, 50))
        button.Bind(wx.EVT_BUTTON, self.Ingresarproceso)
        
        button2 = wx.Button(self.panel, id=wx.ID_ANY, label="Consultar Proceso" ,pos=(900, 150), size=(200, 50))
        button2.Bind(wx.EVT_BUTTON, self.BtnConsultaProceso)
                
        btn_asignar_procesos = wx.Button(self.panel, id=wx.ID_ANY, label="Ident. Nro Proceso\n(Proceso Auto)" ,pos=(900, 200), size=(200, 50))
        btn_asignar_procesos.Bind(wx.EVT_BUTTON, self.onBtn_asignar_procesos)
        
        btn_auto_actuaciones = wx.Button(self.panel, id=wx.ID_ANY, label="Encontrar actuaciones y Enviar Correo\n(Proceso Auto)" ,pos=(900, 250), size=(200, 50))
        btn_auto_actuaciones.Bind(wx.EVT_BUTTON, self.onBtn_encontrar_actuaciones)
        
        btn_actualizar_proceso = wx.Button(self.panel, id=wx.ID_ANY, label="Actualizar Proceso" ,pos=(900, 300), size=(200, 50))
        btn_actualizar_proceso.Bind(wx.EVT_BUTTON, self.onBtn_actualizar_proceso)
        
        btn_usuarios = wx.Button(self.panel, id=wx.ID_ANY, label="Usuarios" ,pos=(900, 350), size=(200, 50))
        btn_usuarios.Bind(wx.EVT_BUTTON, self.onBtn_usuarios)
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        
        pub.subscribe(self.myListener, "frameListener")
        
        # Ask user to login
        dlg = LoginDialog()
        dlg.ShowModal()

    def myListener(self, message, arg2=None):
        """
        Show the frame
        """
        self.Show()
    
    
    #-------------Button Functions-----------------#
    def Ingresarproceso(self, event):
        secondWindow = ww_Ingresar_Proceso(parent=self.panel)
        secondWindow.Show()

    def BtnConsultaProceso(self, event): 
        consultawindow=ww_Consultar_Proceso(parent=self.panel)
        consultawindow.Show()

    def onBtn_asignar_procesos(self, event):
        asignar_nro_proceso()

    def onBtn_encontrar_actuaciones(self, event):
        encontrar_actuaciones()
        
    def onBtn_actualizar_proceso(self, event):
        ww_actualizar_proceso(parent=self.panel).Show()
    
    def onBtn_usuarios(self, event):
        ww_Users(parent=self.panel).Show()
                
    #-------------Button Functions-----------------#    
        
        
class ww_actualizar_proceso(wx.Frame):
    
    
    def __init__(self,parent):
        wx.Frame.__init__(self,parent, -1,'Actualizar Proceso', size=(600,750))  
        self.Centre()
        try:
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
             #   self, -1, bmp1, (0, 0))
            self.panel=wx.Panel(self)
            self.panel.SetBackgroundColour(wx.Colour('white'))
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
            
        wb_database=openpyxl.load_workbook('Database-Process.xlsx')
        self.db_sheet=wb_database['Hoja1']
        
        self.wb_actuaciones=openpyxl.load_workbook('Database-Actuaciones.xlsx')
        self.act_sheet=self.wb_actuaciones['Actuaciones']
        
        estados=make_others_list()
        self.estado_abierto=estados[5][1]
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(20, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        bold_font= wx.Font(70, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        
        self.info_clients=get_clients_info(self.db_sheet,self.estado_abierto)
        list_clients_open=self.info_clients[0]
        
        self.id_clients=self.info_clients[1]
        self.estado_clients=self.info_clients[2]
        
        self.lblactualizar_procesos=wx.StaticText(self.panel, label='Actualizar Procesos')
        self.lblclientes_abiertos=wx.StaticText(self.panel, label='Clientes Con\nProcesos Abiertos')
        self.lblprocesos_abiertos=wx.StaticText(self.panel, label='Procesos Abiertos')
        self.lblactuaciones_pend=wx.StaticText(self.panel, label='Actuaciones\nPendientes')
        self.lblactuacion=wx.StaticText(self.panel, label='Actuacion')
        self.lblrptaactuacion=wx.StaticText(self.panel, label='')
        self.lbldescripcion=wx.StaticText(self.panel, label='\nDescripcion\n\n')
        self.txtdescripcion=wx.TextCtrl(self.panel ,style=wx.TE_MULTILINE | wx.TE_READONLY)
        self.lblfecha_actuacion=wx.StaticText(self.panel, label='Fecha Actuacion')
        self.lblrptafecha_actuacion=wx.StaticText(self.panel, label='')
        self.lblfecha_fin_termino=wx.StaticText(self.panel, label='Fecha Fin Termino')
        self.lblrptafecha_fin_termino=wx.StaticText(self.panel, label='')
        self.lblactuacion_relacionada=wx.StaticText(self.panel, label='¿La actuacion esta relacionada con algunas de las actuaciones anteriores?')
        self.lblseleccione_actuacion_relacionada=wx.StaticText(self.panel, label='Seleccione Actuacion a Relacionar')
        self.lblactuacion_propia=wx.StaticText(self.panel, label='¿La actuacion es Propia?')
        self.lblinfo_adicional=wx.StaticText(self.panel, label='Informacion Adicional')
        self.lblestrategia=wx.StaticText(self.panel, label='Estrategia')
        self.lblfecha_limite=wx.StaticText(self.panel, label='Fecha Limite')
        self.lblestado=wx.StaticText(self.panel, label='Estado')
        self.lbrptaactuacion=wx.StaticText(self.panel, label='')
        self.lbrptadescripcion=wx.StaticText(self.panel, label='')
        self.lbrptafecha_actuacion=wx.StaticText(self.panel, label='')
        self.lbrptafecha_fin_termino=wx.StaticText(self.panel, label='')
        self.comboclientes_abiertos=wx.ComboBox(self.panel,value='', choices=list_clients_open)
        self.comboprocesos_abiertos=wx.ComboBox(self.panel,value='', choices=[])
        self.comboactuaciones_pend=wx.ComboBox(self.panel,value='', choices=[])
        self.comboactuacion_relacionada=wx.ComboBox(self.panel, value='', choices=[])
        self.comboestado=wx.ComboBox(self.panel,value='', choices=estado_choices)
        self.txtinfo_adicional=wx.TextCtrl(self.panel,style=wx.TE_MULTILINE)
        self.checkactuacion_propia_si=wx.CheckBox(self.panel, label= "Si")
        self.checkactuacion_propia_no=wx.CheckBox(self.panel, label= "No")
        self.checkactuacion_relacionada_si=wx.CheckBox(self.panel, label= "Si")
        self.checkactuacion_relacionada_no=wx.CheckBox(self.panel, label= "No")
        
        
        btn_actualizar = wx.Button(self.panel, label="Actualizar",size=(-1,-1))
        btn_cancelar=wx.Button(self.panel, label="Cancelar",size=(-1,-1))

        
        self.lblactualizar_procesos.SetFont(title_font)
        self.lblactuacion.SetFont(bold_font)
        self.lbldescripcion.SetFont(bold_font)
        self.lblfecha_actuacion.SetFont(bold_font)
        self.lblfecha_fin_termino.SetFont(bold_font)
        
        
        self.fgs.Add(self.lblactualizar_procesos, pos=(1,2),span=(1,3), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblclientes_abiertos, pos=(3,1),span=(1,1), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblprocesos_abiertos, pos=(5,1),span=(1,1), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblactuaciones_pend, pos=(7,1),span=(1,1), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblactuacion, pos=(10,1),span=(1,1), flag= wx.ALL, border=2)
        self.fgs.Add(self.lblrptaactuacion, pos=(10,2),span=(1,1), flag= wx.ALL, border=2)
        self.fgs.Add(self.lbldescripcion, pos=(11,1),span=(3,1), flag= wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=2)
        self.fgs.Add(self.txtdescripcion, pos=(11,2),span=(3,4), flag= wx.ALL | wx.EXPAND, border=2)
        self.fgs.Add(self.lblfecha_actuacion, pos=(14,1),span=(1,1), flag= wx.ALL, border=2)
        self.fgs.Add(self.lblrptafecha_actuacion, pos=(14,2),span=(1,1), flag= wx.ALL, border=2)
        self.fgs.Add(self.lblfecha_fin_termino, pos=(15,1),span=(1,1), flag= wx.ALL, border=2)
        self.fgs.Add(self.lblrptafecha_fin_termino, pos=(15,2),span=(1,1), flag= wx.ALL, border=2)
        self.fgs.Add(self.lblactuacion_relacionada, pos=(16,1),span=(2,3), flag= wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=5)
        self.fgs.Add(self.lblseleccione_actuacion_relacionada, pos=(18,1),span=(1,2), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblactuacion_propia, pos=(20,1),span=(2,1), flag= wx.ALL | wx.ALIGN_CENTER_VERTICAL, border=5)
        self.fgs.Add(self.lblinfo_adicional, pos=(24,1),span=(1,3), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblestrategia, pos=(29,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblfecha_limite, pos=(29,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblestado, pos=(29,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.comboclientes_abiertos, pos=(3,2),span=(1,1), flag= wx.ALL, border=0)
        self.fgs.Add(self.comboprocesos_abiertos, pos=(5,2),span=(1,1), flag= wx.ALL, border=0)
        self.fgs.Add(self.comboactuaciones_pend, pos=(7,2),span=(1,3), flag= wx.ALL| wx.EXPAND, border=0)
        self.fgs.Add(self.comboactuacion_relacionada, pos=(18,3),span=(1,3), flag= wx.ALL| wx.EXPAND, border=5)
        self.fgs.Add(self.comboestado, pos=(30,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtinfo_adicional, pos=(25,1),span=(4,4), flag= wx.ALL | wx.EXPAND, border=0)
        self.fgs.Add(self.checkactuacion_propia_si, pos=(20,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.checkactuacion_propia_no, pos=(21,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.checkactuacion_relacionada_si, pos=(16,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.checkactuacion_relacionada_no, pos=(17,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(btn_actualizar, pos=(29,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(btn_cancelar, pos=(30,4),span=(1,1), flag= wx.ALL, border=5)

        self.checkactuacion_propia_si.Bind(wx.EVT_CHECKBOX, self.oncheckactuacion_propia_si)
        self.checkactuacion_propia_no.Bind(wx.EVT_CHECKBOX, self.oncheckactuacion_propia_no)
        self.checkactuacion_relacionada_si.Bind(wx.EVT_CHECKBOX, self.oncheckactuacion_relacionada_si)
        self.checkactuacion_relacionada_no.Bind(wx.EVT_CHECKBOX, self.oncheckactuacion_relacionada_no)
        
               
        self.comboclientes_abiertos.Bind(wx.EVT_COMBOBOX, self.get_open_client_process)
        self.comboprocesos_abiertos.Bind(wx.EVT_COMBOBOX, self.get_open_actuaciones_process)
        self.comboactuaciones_pend.Bind(wx.EVT_COMBOBOX, self.get_act_info)

        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer)
        
        self.lblseleccione_actuacion_relacionada.Hide()
        self.comboactuacion_relacionada.Hide()
        self.lblactuacion_propia.Hide()
        self.checkactuacion_propia_no.Hide()
        self.checkactuacion_propia_si.Hide()
        
        
        btn_actualizar.Bind(wx.EVT_BUTTON, self.Onactualizar)
        btn_cancelar.Bind(wx.EVT_BUTTON, self.Oncancelar)

    def oncheckactuacion_propia_si(self,event):
        if self.checkactuacion_propia_no.IsChecked():
            self.checkactuacion_propia_no.SetValue(False)
    
    def oncheckactuacion_propia_no(self,event):

        if self.checkactuacion_propia_si.IsChecked():
            self.checkactuacion_propia_si.SetValue(False)
                          
    
    def get_open_client_process(self, event):
        
        cliente = self.comboclientes_abiertos.GetValue()
        self.info_clients2= get_client_process_open(self.db_sheet,self.estado_abierto,self.id_clients, self.estado_clients, cliente)
        procesos_abiertos=self.info_clients2[0]
        self.radicados_ini=self.info_clients2[1]
        
        self.comboprocesos_abiertos.Clear()
        self.comboprocesos_abiertos.AppendItems(procesos_abiertos)
        
    def get_open_actuaciones_process(self,event):
        
        radicado_ini = self.comboprocesos_abiertos.GetValue()
        self.info_actuaciones= get_actuaciones_process_open(self.act_sheet,self.estado_abierto,radicado_ini)
        actuaciones_abiertas= self.info_actuaciones[0]
        
        grupos_actuaciones=self.info_actuaciones[2]
        self.comboactuacion_relacionada.Clear()
        self.comboactuacion_relacionada.AppendItems(grupos_actuaciones)
        
        self.index_actuaciones_abiertas=self.info_actuaciones[1]
        self.comboactuaciones_pend.Clear()
        self.comboactuaciones_pend.AppendItems(actuaciones_abiertas)
    
    def get_act_info(self,event):
        index_opc_selec=self.comboactuaciones_pend.GetSelection()
        self.index_actuacion_selec=self.index_actuaciones_abiertas[index_opc_selec]
        
        actuacion=self.comboactuaciones_pend.GetValue()
        descripcion=self.act_sheet.cell(row=(self.index_actuaciones_abiertas[index_opc_selec]+1),column=col_anotacion).value
        fecha_actuacion=self.act_sheet.cell(row=(self.index_actuaciones_abiertas[index_opc_selec]+1),column=col_fecha_actuacion).value
        fecha_fin=self.act_sheet.cell(row=(self.index_actuaciones_abiertas[index_opc_selec]+1),column=col_fecha_fin_termino).value
        
        self.lblrptaactuacion.SetLabel(actuacion)
        
        if descripcion !=None:
             self.txtdescripcion.SetValue(descripcion)   
        
        if fecha_actuacion !=None:
            self.lblrptafecha_actuacion.SetLabel(fecha_actuacion)
        
        if fecha_fin !=None:
            self.lblrptafecha_fin_termino.SetLabel(fecha_fin)
            
        if 'memorial' in actuacion.lower():
            self.memorial=True
            self.lblactuacion_propia.Show()
            self.checkactuacion_propia_si.Show()
            self.checkactuacion_propia_no.Show()
        
        else:
            self.memorial=False
            self.lblactuacion_propia.Hide()
            self.checkactuacion_propia_si.Hide()
            self.checkactuacion_propia_no.Hide()
        
    
    def oncheckactuacion_relacionada_si(self,event):
        self.lblseleccione_actuacion_relacionada.Show()
        self.comboactuacion_relacionada.Show()
        
        
        if self.checkactuacion_relacionada_no.IsChecked():
            self.checkactuacion_relacionada_no.SetValue(False)
            
    
    def oncheckactuacion_relacionada_no(self,event):
        self.lblseleccione_actuacion_relacionada.Hide()
        self.comboactuacion_relacionada.Hide()

        if self.checkactuacion_relacionada_si.IsChecked():
            self.checkactuacion_relacionada_si.SetValue(False)
        
    
    def Onactualizar(self,event):
        actuacion=self.comboactuacion_relacionada.GetValue()
        dic_act_grupo=self.info_actuaciones[3]
        grupos_actuaciones=self.info_actuaciones[4]
        estado=self.comboestado.GetValue()
        
        if self.checkactuacion_relacionada_si.GetValue()==True:
            self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_principal).value='No'
            self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_grupo).value=dic_act_grupo[actuacion]
        elif self.checkactuacion_relacionada_no.GetValue()==True:
            self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_principal).value='Si'
            if grupos_actuaciones:
                max_group=max(grupos_actuaciones)
            else:
                max_group=0
            self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_grupo).value=(max_group+1)
        
        if self.memorial==True:
            if self.checkactuacion_propia_si.IsChecked():
                self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_actuacion_propia).value='Si'
            elif self.checkactuacion_propia_no.IsChecked():
                self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_actuacion_propia).value='No'
        else:
            self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_actuacion_propia).value='N/A'
        
        if estado==estado_choices[1]:
            self.act_sheet.cell(row=(self.index_actuacion_selec+1),column=col_estado).value=estado
        
        
        self.wb_actuaciones.save('Database-Actuaciones.xlsx')
        
        self.comboclientes_abiertos.SetValue("")
        self.comboprocesos_abiertos.SetValue("")
        self.comboactuaciones_pend.SetValue("")
        self.comboactuacion_relacionada.SetValue("")
        self.comboestado.SetValue("")
        self.lblrptaactuacion.SetLabel("")
        self.lblrptafecha_actuacion.SetLabel("")
        self.lblrptafecha_fin_termino.SetLabel("")
        self.txtdescripcion.SetLabel("")
        self.checkactuacion_relacionada_no.SetValue(False)
        self.checkactuacion_relacionada_si.SetValue(False)
        self.checkactuacion_propia_no.SetValue(False)
        self.checkactuacion_propia_si.SetValue(False)
        self.lblseleccione_actuacion_relacionada.Hide()
        self.comboactuacion_relacionada.Hide()
        self.lblactuacion_propia.Hide()
        self.checkactuacion_propia_no.Hide()
        self.checkactuacion_propia_si.Hide()        
        
        
    
    def Oncancelar(self,event):
        self.Destroy()
    
            
        
class ww_Ingresar_Proceso(wx.Frame):
   
    
    def __init__(self,parent):
        
        wx.Frame.__init__(self,parent, -1,'Ingresar Proceso', size=(880,570))
        ciudades_entidades=make_cities_entities_dictionary()
        self.other_lists=make_others_list()
        self.ciudad='MEDELLIN '
        
        try:
            
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
                #self, -1, bmp1, (0, 0)
            self.panel=wx.Panel(self)
            self.panel.SetBackgroundColour(wx.Colour('white'))
            
            
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        title_font= wx.Font(25, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        
        self.fgs= wx.GridBagSizer(0,0)
        
        self.SetIcon(ico)
        
        self.lbltitle =wx.StaticText(self.panel, label='Nuevo Proceso')
        self.lbltitle.SetFont(title_font)
        self.lbltitle.SetBackgroundColour('white')
        self.fgs.Add(self.lbltitle,pos=(0,4),span=(1,3), flag=wx.ALL | wx.ALIGN_CENTER, border=5)
        
        ################################ CIUDAD ############################################
        self.lblciudad = wx.StaticText(self.panel, label="Ciudad:")
        self.lblciudad.SetBackgroundColour("white")
        self.fgs.Add(self.lblciudad,pos=(2,1),span=(1,1), flag= wx.ALL, border=5)
        self.Ciudad=wx.ComboBox(self.panel, value='Ciudad', choices=ciudades_entidades[1])
        self.Ciudad.Bind(wx.EVT_COMBOBOX, self.get_entidades)
        self.fgs.Add(self.Ciudad,pos=(2,2),span=(1,2), flag= wx.ALL , border=5)
        ################################ CIUDAD ############################################
    
        ################################ ENTIDAD ############################################
        self.lblentidad = wx.StaticText(self.panel, label="Entidad:")
        self.lblentidad.SetBackgroundColour("white")
        self.fgs.Add(self.lblentidad,pos=(3,1),span=(1,1), flag= wx.ALL, border=5)
        self.Entidades=wx.ComboBox(self.panel, choices=[""],size=(520,-1))
        self.fgs.Add(self.Entidades,pos=(3,2),span=(1,5), flag=wx.ALL , border=5)
        ################################ ENTIDAD ############################################

        ################################ JURISDICCION ############################################
        self.lbljurisdiccion = wx.StaticText(self.panel, label="Jurisdicción:")
        self.lbljurisdiccion.SetBackgroundColour("white")
        self.fgs.Add(self.lbljurisdiccion, pos=(4,1),span=(1,1), flag= wx.ALL, border=5)
        self.Jurisdi = wx.TextCtrl(self.panel)
        self.fgs.Add(self.Jurisdi, pos=(4,2),span=(1,2), flag= wx.ALL | wx.EXPAND , border=5)
        ################################ JURISDICCION ############################################
        
        ################################ TIPO_SUJETO ############################################
        self.lbltipo_sujeto = wx.StaticText(self.panel, label="Tipo Sujeto \n Cliente:")
        self.lbltipo_sujeto.SetBackgroundColour("white")
        self.fgs.Add(self.lbltipo_sujeto, pos=(5,1),span=(1,1), flag= wx.ALL | wx.EXPAND  , border=5)
        self.Tipsuj = wx.ComboBox(self.panel ,value=self.other_lists[0][0], choices=self.other_lists[0])
        self.Tipsuj.Bind(wx.EVT_COMBOBOX, self.get_tercero)
        self.fgs.Add(self.Tipsuj, pos=(5,2),span=(1,2), flag= wx.ALL | wx.EXPAND , border=5)
        ################################ TIPO SUJETO ############################################
        
        ################################ TIPO PROCESO ############################################
        self.lbltipo_proceso=wx.StaticText(self.panel, label='Tipo Proceso')
        self.lbltipo_proceso.SetBackgroundColour("white")
        self.fgs.Add(self.lbltipo_proceso , pos=(6,1),span=(1,1), flag= wx.ALL, border=5)
        self.tipo_proceso = wx.ComboBox(self.panel,value=self.other_lists[2][0], choices=self.other_lists[2])
        self.fgs.Add(self.tipo_proceso , pos=(6,2),span=(1,5), flag= wx.ALL | wx.EXPAND, border=5)
        ################################ TIPO PROCESO ############################################
        
        ################################ SECCION DEMANDANTE ############################################
        self.lbldemandante=wx.StaticText(self.panel, label='Demandante')
        self.lbldemandante.SetBackgroundColour("white")
        self.fgs.Add(self.lbldemandante , pos=(8,1),span=(1,2), flag=wx.ALL | wx.ALIGN_CENTER, border=5)
        self.lbltipo_persona_demandante=wx.StaticText(self.panel, label='Tipo Persona')
        self.lbltipo_persona_demandante.SetBackgroundColour("white")
        self.fgs.Add(self.lbltipo_persona_demandante , pos=(9,1),span=(1,1), flag= wx.ALL, border=5)
        self.tipo_persona_demandante = wx.ComboBox(self.panel,value=self.other_lists[1][0], choices=self.other_lists[1])
        self.tipo_persona_demandante.Bind(wx.EVT_COMBOBOX, self.get_labels_demandante)
        self.fgs.Add(self.tipo_persona_demandante , pos=(9,2),span=(1,1), flag=wx.ALL | wx.EXPAND, border=5)
        self.lblrazon_social_demandante=wx.StaticText(self.panel, label='Razon Social')
        self.lblrazon_social_demandante.SetBackgroundColour("white")
        self.fgs.Add(self.lblrazon_social_demandante , pos=(10,1),span=(1,1), flag= wx.ALL, border=5)
        self.razon_social_demandante = wx.TextCtrl(self.panel)
        self.fgs.Add(self.razon_social_demandante , pos=(10,2),span=(1,1), flag= wx.ALL, border=5)
        self.lblid_demandante=wx.StaticText(self.panel, label='NIT')
        self.lblid_demandante.SetBackgroundColour("white")
        self.fgs.Add(self.lblid_demandante , pos=(11,1),span=(1,1), flag= wx.ALL, border=5)
        self.id_demandante = wx.TextCtrl(self.panel)
        self.fgs.Add(self.id_demandante , pos=(11,2),span=(1,1), flag= wx.ALL, border=5)
        ################################ SECCION DEMANDANTE ############################################
        
        ################################ SECCION DEMANDADO ############################################
        self.lbldemandado=wx.StaticText(self.panel, label='Demandado')
        self.lbldemandado.SetBackgroundColour("white")
        self.fgs.Add(self.lbldemandado , pos=(8,4),span=(1,2), flag=wx.ALL | wx.ALIGN_CENTER, border=5)
        self.lbltipo_persona_demandado=wx.StaticText(self.panel, label='Tipo Persona')
        self.lbltipo_persona_demandado.SetBackgroundColour("white")
        self.fgs.Add(self.lbltipo_persona_demandado , pos=(9,4),span=(1,1), flag= wx.ALL, border=5)
        self.tipo_persona_demandado = wx.ComboBox(self.panel, value=self.other_lists[1][0],choices=self.other_lists[1])
        self.tipo_persona_demandado.Bind(wx.EVT_COMBOBOX, self.get_labels_demandado)
        self.fgs.Add(self.tipo_persona_demandado , pos=(9,5),span=(1,1), flag= wx.ALL | wx.EXPAND, border=5)
        self.lblrazon_social_demandado=wx.StaticText(self.panel, label='Razon Social')
        self.lblrazon_social_demandado.SetBackgroundColour("white")
        self.fgs.Add(self.lblrazon_social_demandado , pos=(10,4),span=(1,1), flag= wx.ALL, border=5)
        self.razon_social_demandado = wx.TextCtrl(self.panel)
        self.fgs.Add(self.razon_social_demandado , pos=(10,5),span=(1,1), flag= wx.ALL, border=5)
        self.lblid_demandado=wx.StaticText(self.panel, label='NIT')
        self.lblid_demandado.SetBackgroundColour("white")
        self.fgs.Add(self.lblid_demandado , pos=(11,4),span=(1,1), flag= wx.ALL, border=5)
        self.id_demandado = wx.TextCtrl(self.panel)
        self.fgs.Add(self.id_demandado , pos=(11,5),span=(1,1), flag= wx.ALL, border=5)
        ################################ SECCION DEMANDADO ############################################
        
        ################################ SECCION TERCERO ############################################
        self.lbltercero=wx.StaticText(self.panel, label='Tercero')
        self.lbltercero.SetBackgroundColour("white")
        self.fgs.Add(self.lbltercero , pos=(8,7),span=(1,2), flag=wx.ALL | wx.ALIGN_CENTER, border=5)
        self.lbltipo_persona_tercero=wx.StaticText(self.panel, label='Tipo Persona')
        self.lbltipo_persona_tercero.SetBackgroundColour("white")
        self.fgs.Add(self.lbltipo_persona_tercero , pos=(9,7),span=(1,1), flag= wx.ALL, border=5)
        self.tipo_persona_tercero = wx.ComboBox(self.panel,value=self.other_lists[1][0], choices=self.other_lists[1])
        self.tipo_persona_tercero.Bind(wx.EVT_COMBOBOX, self.get_labels_tercero)
        self.fgs.Add(self.tipo_persona_tercero , pos=(9,8),span=(1,1), flag= wx.ALL | wx.EXPAND, border=5)
        self.lblrazon_social_tercero=wx.StaticText(self.panel, label='Razon Social')
        self.lblrazon_social_tercero.SetBackgroundColour("white")
        self.fgs.Add(self.lblrazon_social_tercero , pos=(10,7),span=(1,1), flag= wx.ALL, border=5)
        self.razon_social_tercero = wx.TextCtrl(self.panel)
        self.fgs.Add(self.razon_social_tercero , pos=(10,8),span=(1,1), flag= wx.ALL, border=5)
        self.lblid_tercero=wx.StaticText(self.panel, label='NIT')
        self.lblid_tercero.SetBackgroundColour("white")
        self.fgs.Add(self.lblid_tercero , pos=(11,7),span=(1,1), flag= wx.ALL, border=5)
        self.id_tercero = wx.TextCtrl(self.panel)
        self.fgs.Add(self.id_tercero , pos=(11,8),span=(1,4), flag= wx.ALL, border=5)
        ################################ SECCION TERCERO ############################################ 

        ################################ CUANTIA ############################################
        self.lblcuantia_ini = wx.StaticText(self.panel, label="Cuantia:")
        self.lblcuantia_ini.SetBackgroundColour("white")
        self.fgs.Add(self.lblcuantia_ini, pos=(13,4),span=(1,1), flag= wx.ALL, border=5)
        self.cuantia_ini = wx.TextCtrl(self.panel)
        self.fgs.Add(self.cuantia_ini, pos=(13,5),span=(1,1), flag= wx.ALL, border=5)        
        ################################ CUANTIA ############################################
        
        ################################ RADICADO ############################################
        self.lblradicado_ini = wx.StaticText(self.panel, label="Radicado Inicial:")
        self.lblradicado_ini.SetBackgroundColour("white")
        self.fgs.Add(self.lblradicado_ini, pos=(13,1),span=(1,1), flag= wx.ALL, border=5)
        self.radicado_ini = wx.TextCtrl(self.panel)
        self.fgs.Add(self.radicado_ini, pos=(13,2),span=(1,1), flag= wx.ALL, border=5)
        ################################ RADICADO ############################################
        
        ################################ FECHA_RADICADO ############################################
        self.lblfecha_rad = wx.StaticText(self.panel, label="Fecha de Radicacion:")
        self.lblfecha_rad.SetBackgroundColour("white")
        self.fgs.Add(self.lblfecha_rad, pos=(14,4),span=(1,1), flag= wx.ALL, border=5)
        self.Fechara = wx.TextCtrl(self.panel)
        self.fgs.Add(self.Fechara, pos=(14,5),span=(1,1), flag= wx.ALL, border=5)        
        ################################ FECHA_RADICADO ############################################
        
        ################################ RESPONSABLE ############################################
        self.lblresponsable = wx.StaticText(self.panel, label="Responsable:")
        self.lblresponsable.SetBackgroundColour("white")
        self.fgs.Add(self.lblresponsable, pos=(14,1),span=(1,1), flag= wx.ALL, border=5)
        self.Responsable = wx.TextCtrl(self.panel)
        self.fgs.Add(self.Responsable, pos=(14,2),span=(1,1), flag= wx.ALL, border=5)
        ################################ RESPONSABLE ############################################
        
        ################################ APODERADO ############################################
        self.lblapoderado_ini = wx.StaticText(self.panel, label="Apoderado:")
        self.lblapoderado_ini.SetBackgroundColour("white")
        self.fgs.Add(self.lblapoderado_ini, pos=(15,4),span=(2,1), flag= wx.ALL, border=5)
        self.apoderado_ini = wx.TextCtrl(self.panel)
        self.fgs.Add(self.apoderado_ini, pos=(15,5),span=(5,1), flag= wx.ALL, border=5)      
        ################################ APODERADO ############################################
        
        ################################ BOTONES ############################################
        btn_crear = wx.Button(self.panel, id=wx.ID_ANY, label="Crear Proceso", size=(200,40))
        self.fgs.Add(btn_crear, pos=(13,7),span=(2,2), flag= wx.ALL, border=0)
        btn_crear.Bind(wx.EVT_BUTTON, self.Crearproceso)
        
        btn_cancelar = wx.Button(self.panel, id=wx.ID_ANY, label="Cancelar",size=(200,40))
        self.fgs.Add(btn_cancelar, pos=(15,7),span=(2,2), flag= wx.ALL, border=0)
        btn_cancelar.Bind(wx.EVT_BUTTON, self.OnCloseWindow)
        ################################ BOTONES ############################################
        
        self.SetBackgroundColour(wx.Colour(100,100,100))
        self.Centre(True)
        self.Show()

        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_CENTER)
        self.panel.SetSizerAndFit(mainSizer)

        self.lbltercero.Hide()
        self.lbltipo_persona_tercero.Hide()
        self.tipo_persona_tercero.Hide()
        self.lblrazon_social_tercero.Hide()
        self.razon_social_tercero.Hide()
        self.lblid_tercero.Hide()
        self.id_tercero.Hide()
        
    def get_entidades(self,event):

        ciudades_entidades=make_cities_entities_dictionary()
        Ciudad = self.Ciudad.GetValue()
        choices=ciudades_entidades[0][Ciudad]
        self.Entidades.Clear()
        self.Entidades.AppendItems(choices)

    def get_tercero(self,event):
        
        TipSuj=self.Tipsuj.GetValue()
        
        if TipSuj != self.other_lists[0][3]:
            self.lbltercero.Hide()
            self.lbltipo_persona_tercero.Hide()
            self.tipo_persona_tercero.Hide()
            self.lblrazon_social_tercero.Hide()
            self.razon_social_tercero.Hide()
            self.lblid_tercero.Hide()
            self.id_tercero.Hide()
        else:
            self.lbltercero.Show()
            self.lbltipo_persona_tercero.Show()
            self.tipo_persona_tercero.Show()
            self.lblrazon_social_tercero.Show()
            self.razon_social_tercero.Show()
            self.lblid_tercero.Show()
            self.id_tercero.Show()
        
    def get_labels_demandante(self,event):
        tipo_persona=self.tipo_persona_demandante.GetValue()
        
        if tipo_persona==self.other_lists[1][1]:
            self.lblrazon_social_demandante.SetLabel('Nombre') 
            self.lblid_demandante.SetLabel('Cedula')
        else:
            self.lblrazon_social_demandante.SetLabel('Razon Social')
            self.lblid_demandante.SetLabel('NIT')
            
    def get_labels_demandado(self,event):        
        tipo_persona=self.tipo_persona_demandado.GetValue()
        
        if tipo_persona==self.other_lists[1][1]:
            self.lblrazon_social_demandado.SetLabel('Nombre') 
            self.lblid_demandado.SetLabel('Cedula')
        else:
            self.lblrazon_social_demandado.SetLabel('Razon Social')
            self.lblid_demandado.SetLabel('NIT')
            
    def get_labels_tercero(self,event):
        tipo_persona=self.tipo_persona_tercero.GetValue()
        
        if tipo_persona==self.other_lists[1][1]:
            self.lblrazon_social_tercero.SetLabel('Nombre') 
            self.lblid_tercero.SetLabel('Cedula')
        else:
            self.lblrazon_social_tercero.SetLabel('Razon Social')
            self.lblid_tercero.SetLabel('NIT')
            
    def OnCloseWindow(self, event):
        self.Destroy()
    
    def Crearproceso(self, event):
        
        
        global col_radicado_ini
        global col_radicado_completo
        global col_fecha_radicacion
        global col_tipo_general_proceso
        global col_tipo_especifico_proceso
        global col_cuantia
        global col_instancia
        global col_responsable
        global col_apoderado
        global col_ciudad
        global col_entidad
        global col_jurisdiccion
        global col_tipo_sujeto_cliente
        global col_tipo_persona_demandante
        global col_razon_social_demandante
        global col_nit_demandate
        global col_tipo_persona_demandado
        global col_razon_social_demandado
        global col_nit_demandado
        global col_tipo_persona_tercero
        global col_razon_social_tercero
        global col_nit_tercero
        global col_nit_cliente
        
        

        smmlv=int(self.other_lists[4])
        limite=40*smmlv
        
        Nproce = 1
        
        while (sheet.cell(row = Nproce, column = 1).value != None) :
          Nproce = Nproce + 1
        sheet.cell(row = Nproce  , column = 1).value = Nproce
        

        Ciudad = self.Ciudad.GetValue()
        sheet.cell(row = Nproce, column = col_ciudad).value = Ciudad
        self.Ciudad.Value=""

        Entidad = self.Entidades.GetValue()
        sheet.cell(row = Nproce, column = col_entidad).value = Entidad
        self.Entidades.Value=""        
        
        Jurisdi = self.Jurisdi.GetValue()
        sheet.cell(row = Nproce, column = col_jurisdiccion).value = Jurisdi
        self.Jurisdi.Value=""
        
        Tipo_sujeto = self.Tipsuj.GetValue()
        sheet.cell(row = Nproce, column = col_tipo_sujeto_cliente).value = Tipo_sujeto 
        self.Tipsuj.Value=self.other_lists[0][0]

        Tipo_persona_demandante= self.tipo_persona_demandante.GetValue()
        sheet.cell(row = Nproce, column = col_tipo_persona_demandante).value = Tipo_persona_demandante
        self.tipo_persona_demandante.Value=self.other_lists[1][0]
        
        Razon_social_demandante=self.razon_social_demandante.GetValue()
        sheet.cell(row = Nproce, column = col_razon_social_demandante).value = Razon_social_demandante
        self.razon_social_demandante.Value=""
        
        Id_demandante=self.id_demandante.GetValue()
        sheet.cell(row = Nproce, column = col_nit_demandate).value = Id_demandante
        self.id_demandante.Value=""        
        
        Tipo_persona_demandado= self.tipo_persona_demandado.GetValue()
        sheet.cell(row = Nproce, column = col_tipo_persona_demandado).value = Tipo_persona_demandado
        self.tipo_persona_demandado.Value=self.other_lists[1][0]
        
        Razon_social_demandado=self.razon_social_demandado.GetValue()
        sheet.cell(row = Nproce, column = col_razon_social_demandado).value = Razon_social_demandado
        self.razon_social_demandado.Value=""
        
        Id_demandado=self.id_demandado.GetValue()
        sheet.cell(row = Nproce, column = col_nit_demandado).value = Id_demandado
        self.id_demandado.Value=""
        
        Tipo_persona_tercero= self.tipo_persona_tercero.GetValue()
        sheet.cell(row = Nproce, column = col_tipo_persona_tercero).value = Tipo_persona_tercero
        self.tipo_persona_tercero.Value=self.other_lists[1][0]
        
        Razon_social_tercero=self.razon_social_tercero.GetValue()
        sheet.cell(row = Nproce, column = col_razon_social_tercero).value = Razon_social_tercero
        self.razon_social_tercero.Value=""
        
        Id_tercero=self.id_tercero.GetValue()
        sheet.cell(row = Nproce, column = col_nit_tercero).value = Id_tercero
        self.id_tercero.Value=""
        
        Cuantia = self.cuantia_ini.GetValue()
        sheet.cell(row = Nproce, column = col_cuantia).value = Cuantia
        self.cuantia_ini.Value=""

        Tipo_proceso=self.tipo_proceso.GetValue()
        sheet.cell(row = Nproce, column = col_tipo_especifico_proceso).value = Tipo_proceso
        self.tipo_proceso.Value=self.other_lists[2][0]
        
        index_tipo_proceso= self.other_lists[2].index(Tipo_proceso)
        self.tipo_proceso_general=self.other_lists[3][index_tipo_proceso]
        sheet.cell(row = Nproce, column = col_tipo_general_proceso).value = self.tipo_proceso_general
        
        if self.tipo_proceso_general=="Declarativo":
            if Tipo_proceso==self.other_lists[2][1]:
                self.instancia="Doble Instancia"
            else:
                self.instancia="Unica Instancia"
        
        elif self.tipo_proceso_general=="De Jurisdicción Voluntaria":
            self.instancia="Unica Instancia"
        else:
            if int(Cuantia) < limite:
                self.instancia="Unica Instancia"
            else:
                self.instancia="Doble Instancia"
         
        sheet.cell(row = Nproce, column = col_instancia).value = self.instancia

        Radicado_ini=self.radicado_ini.GetValue()
        sheet.cell(row = Nproce, column = col_radicado_ini).value = Radicado_ini
        self.radicado_ini.Value=""
        
        Responsable = self.Responsable.GetValue()
        sheet.cell(row = Nproce, column = col_responsable).value = Responsable
        self.Responsable.Value=""
        

        Fechara  = self.Fechara.GetValue()
        sheet.cell(row = Nproce, column = col_fecha_radicacion).value = Fechara
        self.Fechara.Value=""
                
        Apoderado = self.apoderado_ini.GetValue()
        sheet.cell(row = Nproce, column = col_apoderado).value =Apoderado
        self.apoderado_ini.Value=""
        
        if Tipo_sujeto==self.other_lists[0][1]:
            sheet.cell(row = Nproce, column = col_nit_cliente).value=Id_demandante
            sheet.cell(row = Nproce, column = col_nombre_cliente).value=Razon_social_demandante
        elif Tipo_sujeto==self.other_lists[0][2]:
            sheet.cell(row = Nproce, column = col_nit_cliente).value=Id_demandado
            sheet.cell(row = Nproce, column = col_nombre_cliente).value=Razon_social_demandado
        elif Tipo_sujeto==self.other_lists[0][3]:
            sheet.cell(row = Nproce, column = col_nit_cliente).value=Id_tercero
            sheet.cell(row = Nproce, column = col_nombre_cliente).value=Razon_social_tercero
                        
         
        success_msgbox=wx.MessageDialog(None,'Registro añadido - Este mensaje aun no garantiza que nada haya fallado en el proceso de agregar el registro - /n EnConstruccion','Status',wx.OK)
        success_msgbox.ShowModal()
        
        DB.save('Database-Process.xlsx')

class ww_Consultar_Proceso(wx.Frame):
    
    def __init__(self,parent):
        wx.Frame.__init__(self,parent, -1,'Consultar Proceso', size=(1200,700))   
        try:
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            self.panel = wx.StaticBitmap(
                self, -1, bmp1, (0, 0))
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        
        self.lblname1 = wx.StaticText(self.panel, label="Ingrese Numero de Proceso", pos=(830, 250))
        self.lblname1.SetBackgroundColour("white")
        self.numero_consulta=wx.TextCtrl(self.panel, size=(300, -1),pos=(750, 270))

        btn_consultar = wx.Button(self.panel, id=wx.ID_ANY, label="Consultar" ,pos=(840, 300), size=(100, 30))
        btn_consultar.Bind(wx.EVT_BUTTON, self.Consultar_Excel)
        
    def Consultar_Excel(self,event):
        
        numero_consulta=self.numero_consulta.GetValue()
        print(os.getcwd())
        workbook_path=os.getcwd()+'/Procesos/'+ numero_consulta + '.xlsx'
        os.startfile(workbook_path)
        
        
        
 
#################################LOG IN ##########################################   
class LoginDialog(wx.Dialog):
    """
    Class to define login dialog
    """
    #----------------------------------------------------------------------
    def __init__(self):
        """Constructor"""
        wx.Dialog.__init__(self, None, title="Login")
        self.Center()
        # user info
        user_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        user_lbl = wx.StaticText(self, label="Username:")
        user_sizer.Add(user_lbl, 0, wx.ALL|wx.CENTER, 5)
        self.user = wx.TextCtrl(self)
        user_sizer.Add(self.user, 0, wx.ALL, 5)
        
        # pass info
        p_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        p_lbl = wx.StaticText(self, label="Password:")
        p_sizer.Add(p_lbl, 0, wx.ALL|wx.CENTER, 5)
        self.password = wx.TextCtrl(self, style=wx.TE_PASSWORD|wx.TE_PROCESS_ENTER)
        p_sizer.Add(self.password, 0, wx.ALL, 5)
        
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        main_sizer.Add(user_sizer, 0, wx.ALL, 5)
        main_sizer.Add(p_sizer, 0, wx.ALL, 5)
        
        btn = wx.Button(self, label="Login")
        btn.Bind(wx.EVT_BUTTON, self.onLogin)
        main_sizer.Add(btn, 0, wx.ALL|wx.CENTER, 5)
        
        self.SetSizer(main_sizer)
        
    #----------------------------------------------------------------------
    def onLogin(self, event):
        """
        Check credentials and login
        """
        stupid_password = "12"
        user_password = self.password.GetValue()
        if user_password == stupid_password:
            print ("You are now logged in!")
            pub.sendMessage("frameListener", message="show")
            self.Destroy()
        else:
            print ("Username or password is incorrect!")
            
#################################LOG IN ##########################################

class ww_Users(wx.Frame):   
    
    def __init__(self,parent):
        
        wb_users=openpyxl.load_workbook('Usuarios.xlsx')
        sheet_users=wb_users['Usuarios']
        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Usuarios", size=(700, 700),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))  
        
        bold_font= wx.Font(70, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        try:
            
            #image_file = 'CINCO CONSULTORES.jpg'
            #bmp1 = wx.Image(
                #image_file, 
                #wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
                #self, -1, bmp1, (0, 0)
            self.panel=wx.Panel(self)
            self.panel.SetBackgroundColour('white')
            self.Center()
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)

        lista_nombres=[]
        lista_usuarios=[]
        lista_correos=[]
        lista_roles=[]
        lista_tipo_usuarios=[]
        
        rows=[]
        ctrls=[]
        
        for row in sheet_users.iter_rows(max_col=4):
            lbls=[]
            for cell in row:
                lbls.append(cell.value)
            rows.append(lbls)

        
        for cell in sheet_users['A']:
            lista_nombres.append(cell.value)
            
        for cell in sheet_users['B']:
            lista_usuarios.append(cell.value)
        
        for cell in sheet_users['C']:
            lista_correos.append(cell.value)
        
        for cell in sheet_users['D']:
            lista_roles.append(cell.value)
        
        for cell in sheet_users['E']:
            lista_tipo_usuarios.append(cell.value)
        
        for i in range(len(rows)):
            for j in range(len(rows[0])):
                if i == 0:
                    ctrls.append(wx.StaticText(self.panel, label=rows[i][j]))
                    self.fgs.Add(ctrls[-1], pos=((3+i),(2+j)), span=(1,1),flag= wx.ALL | wx.ALIGN_CENTER, border=5 )
                    ctrls[j].SetFont(bold_font)
                else:
                    ctrls.append(wx.StaticText(self.panel, label=rows[i][j]))
                    self.fgs.Add(ctrls[-1], pos=((3+i),(2+j)), span=(1,1),flag= wx.ALL | wx.ALIGN_LEFT , border=5 )
  
        
        
        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer)
       
class MyApp(wx.App):
    def OnInit(self):
        self.frame= MyFrame()
        self.frame.Show()
        return True       
# Run the program     

app=MyApp()
app.MainLoop()
del app