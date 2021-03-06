from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import Select
from copy import copy, deepcopy

from datetime import datetime

import time
import openpyxl
import openpyxl.worksheet.cell_range
from openpyxl.styles import Alignment

import os
import smtplib
from email.message import EmailMessage



timeout=10
timeout2=20

#-- Excel BD Procesos--#
col_radicado_ini=2
col_radicado_completo=3
col_fecha_radicacion=4
col_tipo_general_proceso=5
col_tipo_especifico_proceso=6
col_cuantia=7
col_instancia=8
col_responsable=9
col_apoderado=10
col_ciudad=11
col_entidad=12
col_jurisdiccion=13
col_tipo_sujeto_cliente=14
col_tipo_persona_demandante=15
col_razon_social_demandante=16
col_nit_demandate=17
col_tipo_persona_demandado=18
col_razon_social_demandado=19
col_nit_demandado=20
col_tipo_persona_tercero=21
col_razon_social_tercero=22
col_nit_tercero=23
col_nit_cliente=24
col_nombre_cliente=25
#-- Excel BD Procesos--#

#-- Excel BD Actuaciones--#
col_id_actuacion=1
col_numero_proceso=2
col_radicado_ini_act=3
col_fecha_actuacion=4
col_actuacion=5
col_anotacion=6
col_fecha_ini_termino=7
col_fecha_fin_termino=8
col_fecha_registro=9
col_estado=10
col_grupo=11
col_principal=12
col_actuacion_propia=13
estado_choices=['Abierto','Cerrado']
#-- Excel BD Actuaciones--#


def get_the_web():

    # Specifying incognito mode as you launch your browser[OPTIONAL]
    option = webdriver.ChromeOptions()
    option.add_argument("--incognito")
    # Create new Instance of Chrome in incognito mode
    browser = webdriver.Chrome('.\drivers\chromedriver', options=option)

    
        
    # Go to desired website
    browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
    # Wait 20 seconds for page to load

    try:
        # Wait until the final element [Avatar link] is loaded.
        # Assumption: If Avatar link is loaded, the whole page would be relatively loaded because it is among
        # the last things to be loaded.
        WebDriverWait(browser, timeout).until(EC.visibility_of_element_located((By.CLASS_NAME, "pie")))
    except TimeoutException:
        print("Timed out waiting for page to load")
        browser.quit()
        
        
    return browser


#------------Get all the elements on the web and send the values on the excel file (database)-------------#

def asignar_nro_proceso ():
    global timeout

    global col_radicado_ini
    global col_radicado_completo
    global col_fecha_radicacion
    global col_tipo_general_proceso
    global col_tipo_especifico_proceso
    global col_cuantia
    global col_instancia
    global col_responsable
    global col_apoderado
    global col_ciudad
    global col_entidad
    global col_jurisdiccion
    global col_tipo_sujeto_cliente
    global col_tipo_persona_demandante
    global col_razon_social_demandante
    global col_nit_demandate
    global col_tipo_persona_demandado
    global col_razon_social_demandado
    global col_nit_demandado
    global col_tipo_persona_tercero
    global col_razon_social_tercero
    global col_nit_tercero
    
    browser=get_the_web()
    from get_lists import make_others_list
    other_lists= make_others_list()
    
    #Open Excel workbook
    wb_database=openpyxl.load_workbook('Database-Process.xlsx')
    db_sheet=wb_database['Hoja1']
    

    
    registered_process=len(db_sheet['A'])
    Nproce=1
    
    for i in range(registered_process-1):
        Nproce +=1
        
        if(db_sheet.cell(row=Nproce,column=col_radicado_completo).value == None):
            
            try: 
                WebDriverWait(browser, timeout).until(EC.element_to_be_clickable((By.ID, "ddlCiudad")))                               
            except TimeoutException:
                print("Problema web al seleccionar Ciudad")
                browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
                continue

            dropdown_ciudad = Select(browser.find_element_by_id("ddlCiudad"))
            dropdown_ciudad.select_by_visible_text(db_sheet.cell(row=Nproce,column=col_ciudad).value)
            
            try:
                WebDriverWait(browser, timeout).until(EC.visibility_of_element_located((By.XPATH, "/html/body/form/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/div/table/tbody/tr[3]/td[2]/select/option[2]")))                                
            except TimeoutException:
                print("Problema web al seleccionar la entidad")
                browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
                continue
            
            time.sleep(3) #NUNCA BORRAR ESTE HP SLEEP
            
            dropdown1= Select(browser.find_element_by_id('ddlEntidadEspecialidad'))
            dropdown1.select_by_visible_text(db_sheet.cell(row=Nproce,column=col_entidad).value)
            
            try:
                WebDriverWait(browser, timeout).until(EC.element_to_be_clickable((By.ID, "rblConsulta")))                                
            except TimeoutException:
                print("Problema web al cargar tabla para ingresar parametros de consulta")
                browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
                continue
            
            
            inputElement3 = Select(browser.find_element_by_id("rblConsulta"))
            inputElement3.select_by_visible_text("Consulta por Nombre o Razón social")
        
            try:
                WebDriverWait(browser, timeout).until(EC.element_to_be_clickable((By.ID, "ddlTipoSujeto")))                                
            except TimeoutException:
                print("Problema web al cargar Tipo de Sujeto")
                browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
                continue
            
            
            inputElement4 = Select(browser.find_element_by_id("ddlTipoSujeto"))
            inputElement5 = Select(browser.find_element_by_id("ddlTipoPersona"))
            inputElement6 = browser.find_element_by_id("txtNatural")
            
            if inputElement6.text != None:
                inputElement6.clear()
            
            tipo_sujeto_cliente=db_sheet.cell(row=Nproce,column=col_tipo_sujeto_cliente).value
            
            if tipo_sujeto_cliente==other_lists[0][1]:
                
                inputElement4.select_by_visible_text(other_lists[0][1])
                inputElement5.select_by_visible_text(db_sheet.cell(row=Nproce,column=col_tipo_persona_demandante).value)
                inputElement6.send_keys(db_sheet.cell(row=Nproce,column=col_razon_social_demandante).value)
                
            else:
                
                inputElement4.select_by_visible_text(other_lists[0][2])
                inputElement5.select_by_visible_text(db_sheet.cell(row=Nproce,column=col_tipo_persona_demandado).value)
                inputElement6.send_keys(db_sheet.cell(row=Nproce,column=col_razon_social_demandado).value)
            
            
            inputElementX=browser.find_element_by_id("sliderBehaviorConsultaNom_railElement")
            inputElementX.click()
    
            inputElement7=browser.find_element_by_id("btnConsultaNom")
            inputElement7.click()

            try:
                WebDriverWait(browser, 3).until(EC.visibility_of_element_located((By.ID,'msjError')))
                        
                btncerrar=browser.find_element_by_id('modalError').find_element_by_tag_name('input')
                btncerrar.click()
                inputElement6.clear()
                
                if tipo_sujeto_cliente==other_lists[0][1]:
                
                    inputElement4.select_by_visible_text(other_lists[0][2])
                    inputElement5.select_by_visible_text(db_sheet.cell(row=Nproce,column=col_tipo_persona_demandado).value)
                    inputElement6.send_keys(db_sheet.cell(row=Nproce,column=col_razon_social_demandado).value)
                
                else:
                
                    inputElement4.select_by_visible_text(other_lists[0][2])
                    inputElement5.select_by_visible_text(db_sheet.cell(row=Nproce,column=col_tipo_persona_demandante).value)
                    inputElement6.send_keys(db_sheet.cell(row=Nproce,column=col_razon_social_demandante).value)
                    
                btn_nueva_consulta=browser.find_element_by_id('btnNuevaConsultaNom')
                btn_nueva_consulta.click()
                
                inputElementX=browser.find_element_by_id("sliderBehaviorConsultaNom_railElement")
                inputElementX.click()
                
                inputElement7.click()

            except TimeoutException:
                pass
            
            try:
                WebDriverWait(browser, 3).until(EC.visibility_of_element_located((By.ID,'msjError')))
                btncerrar=browser.find_element_by_id('modalError').find_element_by_tag_name('input')
                btncerrar.click()
                print("La consulta No genero resultados, es decir, el proceso aun no esta en la web")
                browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
                continue
            
            except TimeoutException:
                pass

            try:
                WebDriverWait(browser, timeout2).until(EC.visibility_of_element_located((By.ID,"gvResultadosNum")))
            except TimeoutException:
                print('Problemas al cargar los resultados de la consulta, proceso no asignado')
                browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
                continue
           
            
            #get the web element table in which the processes are contained
            tabla_procesos=browser.find_element_by_id('gvResultadosNum')
            #get all the <td> tags of the table in which the data is contained
            campos_tabla_busqueda=tabla_procesos.find_elements_by_tag_name('td')

    #---------------------------Get all the elements on the web and send the values on the excel file (database)----------------------------#
            
    #---------------------------Assign a process number in the excel file----------------------------#       
    
            #get the number of rows of the table
            cantidad_procesos=len(tabla_procesos.find_elements_by_tag_name('tr'))
    
            lista_numeros_procesos=[]
            lista_fechas_radicacion=[]
                
            #get all the "Fechas de Radicacion", step 7, because dates appear every 7 fields.
            for i in range (2,len(campos_tabla_busqueda),7):
                lista_fechas_radicacion.append(campos_tabla_busqueda[i].text)
            
            #get all the "Numeros de procesos" in the table
            for i in range (cantidad_procesos-1):
            #has to be minus 1 because the heading is included on the list
                try:
                    numero_proceso='gvResultadosNum_lnkActuacionesnNum_'
                    numero_proceso += str(i)   
                    lista_numeros_procesos.append(browser.find_element_by_id(numero_proceso).text)
                except (NoSuchElementException):
                        print('Posiblemente hay 2 paginas de procesos - Esto aun esta en construccion')
            
            #get the "fecha radicacion" from the excel file (database) to compare the dates from the table
            fecha_radicacion=db_sheet.cell(row=Nproce,column=col_fecha_radicacion).value
    
            #assign the number process in the excel file
            if fecha_radicacion in lista_fechas_radicacion:
                if lista_fechas_radicacion.count(fecha_radicacion) >1:
                    print('Hay mas de un proceso con la misma fecha, numero no asignado')
                else:
                    db_sheet.cell(row=Nproce,column=col_radicado_completo).value= lista_numeros_procesos[lista_fechas_radicacion.index(fecha_radicacion)]
                    wb_database.save('Database-Process.xlsx')
                    print('Numero de Proceso Asignado -  OK')
                    create_excel_file (lista_numeros_procesos[lista_fechas_radicacion.index(fecha_radicacion)],flag_browser=0,flag_actuaciones=0,browser=0)
                    browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
            else:
                print('Numero de Proceso no encontrado')
    print('DONE - SIN ERRORES - CESAR PUTO AMO')
    browser.quit()
         
#---------------------------Find and Assign a process number in the excel file----------------------------#
    
def search_process(process_number_given):
    browser=get_the_web()
    wb_database=openpyxl.load_workbook('Database-Process.xlsx')
    db_sheet=wb_database['Hoja1']
    number_process_column=db_sheet['C']
    
    process_numbers=[]
    
    for cell in number_process_column:
        process_numbers.append(cell.value)
    
    fila_proceso=(process_numbers.index(process_number_given)+1)
    
    try: 
        WebDriverWait(browser, timeout).until(EC.element_to_be_clickable((By.ID, "ddlCiudad")))                               
    except TimeoutException:
        print("Problema web al seleccionar Ciudad - Excel no creado")
        browser.quit()
        return
    
    dropdown_ciudad = Select(browser.find_element_by_id("ddlCiudad"))
    dropdown_ciudad.select_by_visible_text(db_sheet.cell(row=fila_proceso,column=col_ciudad).value)
    
    try:
        WebDriverWait(browser, timeout).until(EC.visibility_of_element_located((By.XPATH, "/html/body/form/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/div/table/tbody/tr[3]/td[2]/select/option[2]")))                                
    except TimeoutException:
        print("Problema web al seleccionar la entidad - Excel no creado")
        browser.quit()
        return   
        
    time.sleep(3) #NUNCA BORRAR ESTE HP SLEEP
    
    dropdown1= Select(browser.find_element_by_id('ddlEntidadEspecialidad'))
    dropdown1.select_by_visible_text(db_sheet.cell(row=fila_proceso,column=col_entidad).value)
    
    inputRadicado = browser.find_element_by_id('divNumRadicacion').find_element_by_tag_name('input')
    inputRadicado.send_keys(process_number_given)
        
    try:
        WebDriverWait(browser, timeout).until(EC.element_to_be_clickable((By.ID, "sliderBehaviorNumeroProceso_railElement")))                                
    except TimeoutException:
        print("Problema web al dar click en la barra para iniciar la consulta - Excel no creado")
        browser.quit()
        return
    
    inputElement7=browser.find_element_by_id("sliderBehaviorNumeroProceso_railElement")
    inputElement7.click()

    btnconsulta=browser.find_element_by_xpath('/html/body/form/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/div[3]/table/tbody/tr[4]/td/div[1]/table/tbody/tr[3]/td/input[1]')
    btnconsulta.click()
    
    try:
        WebDriverWait(browser, timeout2).until(EC.visibility_of_element_located((By.CLASS_NAME, "ActuacionesDetalle")))                                
    except TimeoutException:
        print("Problema web al cargar tabla de informacion de Actuaciones - Excel no creado")
        browser.quit()
        return

    return browser

def create_excel_file (process_number_given,flag_browser,flag_actuaciones,browser=0):
    global timeout
    global timeout2
    
    global col_radicado_ini
    global col_radicado_completo
    global col_fecha_radicacion
    global col_tipo_general_proceso
    global col_tipo_especifico_proceso
    global col_cuantia
    global col_instancia
    global col_responsable
    global col_apoderado
    global col_ciudad
    global col_entidad
    global col_jurisdiccion
    global col_tipo_sujeto_cliente
    global col_tipo_persona_demandante
    global col_razon_social_demandante
    global col_nit_demandate
    global col_tipo_persona_demandado
    global col_razon_social_demandado
    global col_nit_demandado
    global col_tipo_persona_tercero
    global col_razon_social_tercero
    global col_nit_tercero
    
    if flag_browser==0:
        browser=search_process(process_number_given)
    else:
        pass
    
    despacho=browser.find_element_by_id('lblJuzgadoActual').text
    ponente=browser.find_element_by_id('lblPonente').text
    tipo=browser.find_element_by_id('lblTipo').text
    clase=browser.find_element_by_id('lblClase').text
    recurso=browser.find_element_by_id('lblRecurso').text
    ubicacion=browser.find_element_by_id('lblUbicacion').text
    demandantes=browser.find_element_by_id('lblNomDemandante').text
    demandados=browser.find_element_by_id('lblNomDemandado').text
    contenido=browser.find_element_by_id('lblContenido').text
    
    
    lista_documentos=[]
    lista_descrip_documentos=[]
    
    try:
        browser.find_element_by_id('rptDocumentos_lbNombreDoc_0')
        tabla_documentos=browser.find_element_by_class_name('DocumentosDetalle')
        cantidad_documentos=tabla_documentos.find_elements_by_tag_name('tr')
        
        for i in range(len(cantidad_documentos)-1):
            nombre_documento='rptDocumentos_lbNombreDoc_'
            descripcion_documento='rptDocumentos_lblDescDoc_'
            
            nombre_documento += str(i)
            lista_documentos.append(browser.find_element_by_id(nombre_documento).text)
            descripcion_documento +=str(i)
            lista_descrip_documentos.append(browser.find_element_by_id(descripcion_documento).text)

    except NoSuchElementException:
        print('No hay documentos Asociados')
        

    tabla_detalle=browser.find_element_by_class_name('ActuacionesDetalle')
    cantidad_actuaciones=len(tabla_detalle.find_elements_by_tag_name('tr'))
    
    lista_fecha_actuaciones=[]
    lista_actuaciones=[]
    lista_anotaciones=[]
    lista_fecha_inicia=[]
    lista_fecha_termina=[]
    lista_fecha_registro=[]

    #we have to substract 1 , due to cantidad_actuaciones is including the header.
    for i in range(cantidad_actuaciones-1):
    
        fecha_actuacion='rptActuaciones_lblFechaActuacion_'
        actuacion='rptActuaciones_lblActuacion_'
        anotacion='rptActuaciones_lblAnotacion_'
        fecha_inicia='rptActuaciones_lblFechaInicio_'
        fecha_termina='rptActuaciones_lblFechaFin_'
        fecha_registro='rptActuaciones_lblFechaRegistro_'
        
        fecha_actuacion += str(i)
        lista_fecha_actuaciones.append(browser.find_element_by_id(fecha_actuacion).text)
        actuacion += str(i)
        lista_actuaciones.append(browser.find_element_by_id(actuacion).text)
        anotacion += str(i)
        lista_anotaciones.append(browser.find_element_by_id(anotacion).text)
        fecha_inicia += str(i)
        lista_fecha_inicia.append(browser.find_element_by_id(fecha_inicia).text)
        fecha_termina += str(i)
        lista_fecha_termina.append(browser.find_element_by_id(fecha_termina).text)
        fecha_registro += str(i)
        lista_fecha_registro.append(browser.find_element_by_id(fecha_registro).text)
    
    
    #Open Excel Workbook
    
    path='./Procesos/Formato_Base.xlsx'
    excel_name_sheet='Formato'
    wb=openpyxl.load_workbook(path)
    main_sheet=wb[excel_name_sheet]
    
    

              
    #fill data in workbook         
              
    main_sheet['C4'].value=despacho
    main_sheet['Z4'].value=ponente
    main_sheet['C9'].value=tipo
    main_sheet['I9'].value=clase
    main_sheet['V9'].value=recurso
    main_sheet['AI9'].value=ubicacion
    main_sheet['C14'].value=demandantes
    main_sheet['Z14'].value=demandados
    main_sheet['C19'].value=contenido
    
    empty_row_doc=26
    
    if lista_documentos:
        
        style_source='C26'
        
        #merged=main_sheet.merged_cells.ranges
        #for i in merged:
        #    i.shift(0,3)        
        main_sheet.insert_rows(empty_row_doc+1, (len(lista_documentos)-1))
        
        
        for i in range(len(lista_documentos)):
            
            main_sheet.cell(row=(empty_row_doc+i),column=3).value=lista_documentos[i]
            main_sheet.cell(row=(empty_row_doc+i),column=26).value=lista_descrip_documentos[i]
            
            main_sheet.cell(row=(empty_row_doc+i), column=3)._style=deepcopy(main_sheet[style_source]._style)
            main_sheet.merge_cells(start_row=empty_row_doc+i, start_column=3, end_row=empty_row_doc+i, end_column=3+22)
            main_sheet.cell(row=(empty_row_doc+i), column=26)._style=deepcopy(main_sheet[style_source]._style)
            main_sheet.merge_cells(start_row=empty_row_doc+i, start_column=26, end_row=empty_row_doc+i, end_column=26+22)        

    else:
        main_sheet.cell(row=(empty_row_doc),column=3).value="No hay Documentos Asociados"
        alignment = Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
        
        
        main_sheet.merge_cells(start_row=empty_row_doc, start_column=3, end_row=empty_row_doc, end_column=3+45)
        main_sheet.cell(row=(empty_row_doc),column=3).alignment=alignment
    
    #define the row number, in which the title "Actuaciones del proceso" is contained
    if not lista_documentos:
        st_row=empty_row_doc+3+(len(lista_documentos))
    else:
        st_row=empty_row_doc+3+(len(lista_documentos)-1)
    
    
    main_sheet.merge_cells(start_row=st_row, start_column=3, end_row=st_row, end_column=3+45)
    main_sheet.merge_cells(start_row=st_row+1, start_column=3, end_row=st_row+1, end_column=3+3)
    main_sheet.merge_cells(start_row=st_row+1, start_column=7, end_row=st_row+1, end_column=7+3)
    main_sheet.merge_cells(start_row=st_row+1, start_column=11, end_row=st_row+1, end_column=11+25)
    main_sheet.merge_cells(start_row=st_row+1, start_column=37, end_row=st_row+1, end_column=37+3)
    main_sheet.merge_cells(start_row=st_row+1, start_column=41, end_row=st_row+1, end_column=41+3)
    main_sheet.merge_cells(start_row=st_row+1, start_column=45, end_row=st_row+1, end_column=45+3)
    
    #Open DB_ Actuaciones to record all data about actuaciones
    wb_db_actuaciones=openpyxl.load_workbook('Database-Actuaciones.xlsx')
    actuaciones_sheet=wb_db_actuaciones['Actuaciones']
    
    
    empty_row_actuaciones=1
    while (actuaciones_sheet.cell(row = empty_row_actuaciones, column = 1).value != None) :
      empty_row_actuaciones += 1
      

    #define the row number, in which the algorithm will start writing the "actuaciones"
    empty_row=st_row+2
    #Define the cell to copy the style
    style_source='C'+str(empty_row)
    
    
        
    for i in range (len(lista_fecha_actuaciones)):
                
        if flag_actuaciones==0:
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_id_actuacion).value=empty_row_actuaciones+i
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_numero_proceso).value=process_number_given
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_fecha_actuacion).value=lista_fecha_actuaciones[i]
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_actuacion).value=lista_actuaciones[i]
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_anotacion).value=lista_anotaciones[i]
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_fecha_ini_termino).value=lista_fecha_inicia[i]
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_fecha_fin_termino).value=lista_fecha_termina[i]
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_fecha_registro).value=lista_fecha_registro[i]
            actuaciones_sheet.cell(row=(empty_row_actuaciones+i),column=col_estado).value=estado_choices[0]
        else:
            pass

        main_sheet.cell(row=(empty_row+i),column=3).value=lista_fecha_actuaciones[i]
        main_sheet.cell(row=(empty_row+i),column=7).value=lista_actuaciones[i]
        main_sheet.cell(row=(empty_row+i),column=11).value=lista_anotaciones[i]
        main_sheet.cell(row=(empty_row+i),column=37).value=lista_fecha_inicia[i]
        main_sheet.cell(row=(empty_row+i),column=41).value=lista_fecha_termina[i]
        main_sheet.cell(row=(empty_row+i),column=45).value=lista_fecha_registro[i]

        main_sheet.row_dimensions[empty_row+i].height = 33
        
        main_sheet.cell(row=(empty_row+i), column=3)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=3, end_row=empty_row+i, end_column=3+3)
        main_sheet.cell(row=(empty_row+i), column=7)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=7, end_row=empty_row+i, end_column=7+3)
        main_sheet.cell(row=(empty_row+i), column=11)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=11, end_row=empty_row+i, end_column=11+25)
        main_sheet.cell(row=(empty_row+i), column=37)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=37, end_row=empty_row+i, end_column=37+3)
        main_sheet.cell(row=(empty_row+i), column=41)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=41, end_row=empty_row+i, end_column=41+3)
        main_sheet.cell(row=(empty_row+i), column=45)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=45, end_row=empty_row+i, end_column=45+3)
        
    
    rd = main_sheet.row_dimensions[st_row] # get dimension for row 3
    rd.height = 18 # value in points, there is no "auto"
    
    rd = main_sheet.row_dimensions[st_row+1]
    rd.height = 27
    
    if flag_browser==0:
        browser.quit()
    else:
        pass
    new_path="./Procesos/" + process_number_given + '.xlsx'
    wb.save(new_path) 
    wb_db_actuaciones.save('Database-Actuaciones.xlsx')
    print('Excel creado exitosamente - OK')
    print('Actuaciones guardadas en la BD - OK')



    
def encontrar_actuaciones():
    
    wb_database=openpyxl.load_workbook('Database-Process.xlsx')
    db_sheet=wb_database['Hoja1']
    
    wb_db_actuaciones=openpyxl.load_workbook('Database-Actuaciones.xlsx')
    actuaciones_sheet=wb_db_actuaciones['Actuaciones']
   
    
    lista_procesos_excel=[]

    for cell in db_sheet['C']:
        if cell.value !=None:
            lista_procesos_excel.append(cell.value)
        
    lista_proceso_actuaciones_excel=[]
    
    for cell in actuaciones_sheet['B']:
        lista_proceso_actuaciones_excel.append(cell.value)
    
    
    for i in range (1,len(lista_procesos_excel)):
        try:
            browser=search_process(lista_procesos_excel[i])
        except:
            browser.quit()
            print('Error se continuara buscando el siguiente proceso')
            continue
    
        fila_radicado_ini=i+1
        radicado_ini=db_sheet.cell(row=fila_radicado_ini, column=col_radicado_ini).value
        tabla_detalle=browser.find_element_by_class_name('ActuacionesDetalle')
        
        #we have to substract 1 , due to cantidad_actuaciones is including the header.
        cantidad_actuaciones_web=len(tabla_detalle.find_elements_by_tag_name('tr'))-1
        cant_actuaciones_excel=lista_proceso_actuaciones_excel.count(lista_procesos_excel[i])

        if cantidad_actuaciones_web>cant_actuaciones_excel:
            create_excel_file(lista_procesos_excel[i],flag_browser=1,flag_actuaciones=1,browser=browser)
            cant_nuevas_actuaciones=cantidad_actuaciones_web-cant_actuaciones_excel
            
            #lista_fecha_actuaciones=[]
            lista_actuaciones_nuevas=[]
            #lista_anotaciones=[]
            #lista_fecha_inicia=[]
            lista_fecha_termina_nuevas=[]
            #lista_fecha_registro=[]

            empty_row_actuaciones=1
            while (actuaciones_sheet.cell(row = empty_row_actuaciones, column = 1).value != None) :
              empty_row_actuaciones += 1
            print(empty_row_actuaciones)
         
            #we have to substract 1 , due to cantidad_actuaciones is including the header.
            for j in range(cant_nuevas_actuaciones):

                fecha_actuacion='rptActuaciones_lblFechaActuacion_'
                actuacion='rptActuaciones_lblActuacion_'
                anotacion='rptActuaciones_lblAnotacion_'
                fecha_inicia='rptActuaciones_lblFechaInicio_'
                fecha_termina='rptActuaciones_lblFechaFin_'
                fecha_registro='rptActuaciones_lblFechaRegistro_'
                
                fecha_actuacion += str(j)
                #lista_fecha_actuaciones.append(browser.find_element_by_id(fecha_actuacion).text)
                actuacion += str(j)
                lista_actuaciones_nuevas.append(browser.find_element_by_id(actuacion).text)
                anotacion += str(j)
                #lista_anotaciones.append(browser.find_element_by_id(anotacion).text)
                fecha_inicia += str(j)
                #lista_fecha_inicia.append(browser.find_element_by_id(fecha_inicia).text)
                fecha_termina += str(j)
                lista_fecha_termina_nuevas.append(browser.find_element_by_id(fecha_termina).text)
                fecha_registro += str(j)
                #lista_fecha_registro.append(browser.find_element_by_id(fecha_registro).text)
            
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_id_actuacion).value=(empty_row_actuaciones+j-1)
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_numero_proceso).value=lista_procesos_excel[i]
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_radicado_ini_act).value=radicado_ini
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_fecha_actuacion).value=browser.find_element_by_id(fecha_actuacion).text
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_actuacion).value=browser.find_element_by_id(actuacion).text
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_anotacion).value=browser.find_element_by_id(anotacion).text
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_fecha_ini_termino).value=browser.find_element_by_id(fecha_inicia).text
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_fecha_fin_termino).value=browser.find_element_by_id(fecha_termina).text
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_fecha_registro).value=browser.find_element_by_id(fecha_registro).text
                actuaciones_sheet.cell(row=(empty_row_actuaciones+j),column=col_estado).value=estado_choices[0]
     
            
            browser.quit()
            send_email('cinconotificaciones@gmail.com',lista_procesos_excel[i],lista_actuaciones_nuevas,lista_fecha_termina_nuevas)
            print('Emails Enviado')
            wb_db_actuaciones.save('Database-Actuaciones.xlsx')
            print('Proceso ' + lista_procesos_excel[i] +' Actualizado')
            
        else:
            browser.quit()
            pass
    print('Proceso Finalizado')

def send_email(receiver,radicado_ini,lista_actuaciones_nuevas,lista_fechafin_nuevas):

    EMAIL_ADDRESS='cinconotificaciones@gmail.com'
    EMAIL_PASSWORD='tydavhmndyxluhbe'
    
    msg = EmailMessage()
    msg['Subject'] = 'Actualizacion Proceso ' + radicado_ini + " // " + str(datetime.today().strftime('%d-%m-%Y'))
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = receiver


    initial_html="""\
    <!DOCTYPE html>
    <html>
    <head>
    <style>
    html{
        font-family: Arial, Helvetica, sans-serif;
    }
    h1{
        color:blue;
    }
    table{
        border-collapse: collapse;
    }
    th,td{
        border: 1px solid black;
    }
    
    </style>
    </head>
        <body>
            <h1>Nueva Actuacion</h1>
            <p>Se informa que el proceso: """+radicado_ini+""" tiene las siguientes actualizaciones.</p>
            <table>
                <tr>
                    <th>Actuacion
                    <th>Fecha Fin de Termino
                </tr>    
    
    """
    
    for i in range (len(lista_actuaciones_nuevas)):
        initial_html+="""<tr>
        <td>"""+lista_actuaciones_nuevas[i]+"""</td>
        <td>"""+lista_fechafin_nuevas[i]+"""</td>
        </tr>"""

    final_html=""" </table>
                    <p>Favor actualizar la informacion de la actuacion</p>
                    <p>Cinco Consultores</p>    
                    </body>
                </html>"""
    
    total_html=initial_html+final_html

    msg.add_alternative(total_html, subtype='html')

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)